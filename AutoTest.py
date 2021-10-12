from openpyxl import load_workbook as LoadExcel
import urllib.parse as urlparse

import logging,os,argparse,sys,difflib

import  serial
import  time
import json,requests

class ExcelToDict:
    def __init__(self,file,title_row=0):
        self.title_row = int(title_row)
        self.ExcelBook = LoadExcel(filename=file)
        self.ExcelDate = {}

    def SheetToList(self,SheetName):
        DataObjList=[]
        titlelist=[]
        if not SheetName in self.ExcelBook.sheetnames:
            raise Exception('SheetName {0} not Existence'.format(SheetName))
        SheetObj = self.ExcelBook[SheetName]
        self.ExcelDate[SheetName]={}

        rows = tuple(SheetObj.rows)
        for row in rows:
            rowsDataObj={}
            if rows.index(row)==self.title_row:
                #这行是title
                for i in row:
                    titlelist.append(i.value)
                continue

            # 读取行里的所有元素
            for i in row:
                rowsDataObj[titlelist[row.index(i)]]=i.value
            DataObjList.append(rowsDataObj)
            self.ExcelDate[SheetName]=DataObjList
        return self.ExcelDate[SheetName]


def SheetDatatoPortlist(PortSheetData):
    Portlist={}
    if type(PortSheetData)!=list:
        raise Exception('config type is {0} , Exception list'.format(type(PortSheetData)))
    for item in PortSheetData:
        Portlist[item['vport']]={}
        parsed = urlparse.urlparse(item['config'])
        Portlist[item['vport']]['type']=parsed.scheme
        Portlist[item['vport']]['config']={}
        
        if parsed.scheme=='uart':
            #print(urlparse.parse_qs(parsed.query))
            Portlist[item['vport']]['config'].update(urlparse.parse_qs(parsed.query))
        elif parsed.scheme=='tcp':
            Portlist[item['vport']]['config']['ip']=parsed.netloc
        elif parsed.scheme=='udp':
            Portlist[item['vport']]['config']['ip']=parsed.netloc
        elif parsed.scheme=='http':
            Portlist[item['vport']]['config']['url']=item['config']
        print(Portlist)
    return Portlist



def TestOneCmd(cmdobj,rcvdfunc,sendfunc):
    sendfunc(cmdobj["send"])
    #logger.debug('send: "%s" OK', cmdobj["send"].replace('\r','<CR>').replace('\n','<LF>'))

    relrcvd = rcvdfunc(cmdobj["rcvd"],cmdobj["timeout"])
    #logger.debug('rcvd "%s" ',relrcvd)
    cmpRes=True
    CmdTimeoutError=False

    if len(relrcvd)==0:
        #接收超时了
        cmpRes=False
        CmdTimeoutError=True
    else:
        relrcvd = relrcvd.replace('\n','')
        for i in range(0,len(cmdobj["rcvd"])):
            if relrcvd[i] != cmdobj["rcvd"][i]:
                cmpRes=False
                break
    
    relrcvd = relrcvd.replace('\r','<CR>').replace('\n','<LF>')

    if cmpRes == True:
        #logger.debug('rcvd "%s" OK',relrcvd)
        time.sleep(cmdobj["wait"])
    else:
        if( cmdobj.has_key("ignor_error") == True) and (cmdobj["ignor_error"] == True ):
            #logger.warning('rcvd "%s",but except "%s",ignor error',relrcvd,cmdobj["rcvd"].replace('\r','<CR>').replace('\n','<LF>'))
            return True
        if 	CmdTimeoutError==False:
            #logger.error('rcvd: "%s",but except "%s"',relrcvd,cmdobj["rcvd"].replace('\r','<CR>').replace('\n','<LF>'))
            CmdTimeoutError=False
        else:
            CmdTimeoutError=False
            #logger.error('rcvd: "%s",timeout %d',relrcvd,cmdobj["timeout"])
        return False
    return True


def openlogfile(logfilepath):
    global LogF
    global logger
    logger=logging.getLogger()
    logger.setLevel(logging.DEBUG)
    LogF = logging.FileHandler(logfilepath, mode='w+')
    LogF.setLevel(logging.DEBUG)
    LogF.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    logger.addHandler(LogF)

def app_exit(id):
    logmsg='exit {0}'.format(id)
    logger.info(logmsg)
    print(logmsg)
    input('按任意按键退出...')
    sys.exit(1)

def serial_test(portconfig,cmd):
    errorflag=False
    config=portconfig['config']
    if not 'Handle' in portconfig:
        Handle=serial.Serial(str(config['port'][0]))
        Handle.baudrate=115200
        Handle.bytesize=int(config['databits'][0])
        Handle.stopbits=int(config['stopbits'][0])
        Handle.parity=serial.PARITY_NONE
        Handle.xonxoff=False
        Handle.rtscts=False
        Handle.dsrdtr=False

        portconfig['Handle']=Handle
        print('open {0} baud {1}'.format(str(config['port'][0]),int(config['baud'][0])))
        del Handle

    Handle=portconfig['Handle']
    Handle.timeout=float(cmd['timeout']) 
    Handle.inter_byte_timeout=float(cmd['timeout'])/20

    # 如果字符串中有<HEX>标记,表示后面的数据发送使用HEX
    SendByteData=cmd['send'].replace('<CR>','\r').replace('<LF>','\n').encode()
    
    Handle.flush()
    #Handle.reset_input_buffer()
    #Handle.reset_output_buffer()
    Handle.write(SendByteData)
    logmsg='send to   {0}:{1}'.format(cmd['port'],SendByteData)
    logger.info(logmsg)
    print(logmsg)

    if cmd['rcvd']==None
    ExceptRcvdByteData=cmd['rcvd'].replace('<CR>','\r').replace('<LF>','\n').encode()

    RcvdMatchLen=0;timecnt=0;currentime=time.time();RcvdByteData=b''
    while not (timecnt>float(cmd['timeout']) or RcvdMatchLen==len(ExceptRcvdByteData)) :
        timecnt=time.time()-currentime
        ByteData=Handle.read()
        if ByteData==b'':
            #接收1个字节超时
            continue
        else:
            RcvdByteData +=ByteData
            if ByteData == bytes([ExceptRcvdByteData[RcvdMatchLen]]):
                RcvdMatchLen+=1
            else:
                RcvdMatchLen=0

    if RcvdMatchLen==len(ExceptRcvdByteData):
        logmsg='rcvd from {0}:{1}'.format(cmd['port'],RcvdByteData)
        print(logmsg)
        logger.info(logmsg)
    elif timecnt > float(cmd['timeout']):
        logmsg='rcvd from {0}:timeout({1})s,current rcvd :{2},Except :{3}'.format(cmd['port'],float(cmd['timeout']),RcvdByteData,ExceptRcvdByteData)
        print(logmsg)
        logger.error(logmsg)
        errorflag=True
    else:
        logmsg='rcvd from {0}:timeout({1})s,current rcvd :{2},Except :{3} RcvdMatchLen :{4} '.format(cmd['port'],float(cmd['timeout']),RcvdByteData,ExceptRcvdByteData,RcvdMatchLen)
        print(logmsg)
        logger.error(logmsg)
        errorflag=True

    if errorflag==True:
        if(cmd['ignore_error']=='Y' or cmd['ignore_error']=='y'):
            print('ignore error ')
        else:
            app_exit(-1)
    del Handle


def http_test(portconfig,cmd):
    if type(cmd['send'])==str:
        cmd['send']=json.loads(cmd['send'])
    methed=cmd['send']['methed']
    url=portconfig['config']['url'] + cmd['send']['url']
    payload=cmd['send']['payload']
    headers=cmd['send']['headers']
    logmsg='send to {0}: url={1}:{2} headers={3} payload={4} '.format(cmd['port'],methed,url,headers,payload)
    logger.info(logmsg)
    print(logmsg)
    try:
        logging.getLogger("requests").setLevel(logging.WARNING)
        response = requests.request(methed, url, headers=headers, data=payload,timeout=float(cmd['timeout']))
        if cmd['rcvd'] == None:
            print("Don't need check rcvd.")
        elif cmd['rcvd'] in response.text:
            logmsg='rcvd from {0}: {1}'.format(cmd['port'],response.text.encode())
            logger.info(logmsg)
        else:
            logmsg='rcvd from {0}: {1},but except {2}'.format(cmd['port'],response.text.encode(),cmd['rcvd'].encode())
            logger.info(logmsg)
    except requests.RequestException as err:
        logmsg = str(err)
        logger.error(logmsg)

    print(logmsg)

    return

# 为了print的时候刷新缓冲
class ForceRefresh(object):
   def __init__(self, stream):
       self.stream = stream
   def write(self, data):
       self.stream.write(data)
       self.stream.flush()
   def __getattr__(self, attr):
       return getattr(self.stream, attr)


def test_sheet(Cmdlist,PortList):
    for iter in Cmdlist:
        if not iter['port'] in PortList:
            raise Exception('maybe excel port info err? no found cmd sheet {0} in port sheet'.format(iter['port']))

        if PortList[iter['port']]['type']=='uart':
            serial_test(PortList[iter['port']],iter)
        elif PortList[iter['port']]['type']=='tcp':
            print('send to',PortList[iter['port']]['type'],'msg:',iter['send'])
        elif PortList[iter['port']]['type']=='udp':
            print('send to',PortList[iter['port']]['type'],'msg:',iter['send'])
        elif PortList[iter['port']]['type']=='http':
            http_test(PortList[iter['port']],iter)

        logmsg='wait {0}'.format(iter['wait'])
        print(logmsg)
        logger.info(logmsg)
        if str(iter['wait']) != '0':
            time.sleep(float(iter['wait']))
    #for end
    return

if __name__ == '__main__':
    sys.stdout = ForceRefresh(sys.stdout)
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '-f',
        '--file',
        nargs='?',
        default='.\\demo\\CMD.xlsx',# for debug
        type=str,
        help='input file(.xlsx)')
    parser.add_argument(
        '-o',
        '--log',
        nargs='?',
        default='log_'+time.strftime('%Y%m%d-%H%M%S',time.localtime(time.time()))+'.log',
        type=str,
        help='output logfile(.txt)')
    #parser.add_argument('-v', '--verbosity', help='increase output verbosity')

    args = parser.parse_args()
    openlogfile(args.log)

    if args.file==None:
        logger.info('input file is None')
        logger.info('{0}'.format(parser.print_help()))
        sys.exit(1)
    
    Excelfile=args.file
    logger.info('open '+Excelfile)
    ExcelObj=ExcelToDict(Excelfile)

    PortList=SheetDatatoPortlist(ExcelObj.SheetToList('Port'))
    CmdSheetList=[]
    for SheetName in ExcelObj.ExcelBook.sheetnames:
        if SheetName != 'Port':
            obj={}
            obj['name']=SheetName
            obj['list']=ExcelObj.SheetToList(SheetName)
            CmdSheetList.append(obj)
    loopnum=0
    while True:
        loopnum=loopnum+1

        logmsg='loop num {0}'.format(loopnum)
        logger.info(logmsg)
        print(logmsg)

        for  Cmdlist in CmdSheetList:
            logmsg='Test Sheet :{0}'.format(Cmdlist['name'])
            logger.info(logmsg)
            print(logmsg)
            test_sheet(Cmdlist['list'],PortList)
        # for end
    #while end

    for iter in PortList:
        if 'PortHandle' in iter:
            iter['PortHandle'].close()
            print('close {0} '.format(str(iter['port'][0])))

    LogF.close()


